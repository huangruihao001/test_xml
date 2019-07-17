import openpyxl
from openpyxl import Workbook

class ReadExcel():


    def __init__(self, workbook_path, sheet_name):
        self.workbook_path = workbook_path # excel路径
        self.sheet_name = sheet_name # 表格名称
        self.xml_workbook = openpyxl.load_workbook(self.workbook_path)
        self.xml_sheet = self.xml_workbook[self.sheet_name]
        self.order = self.xml_sheet["A2"].value # 主订单号
        self.Eorder = self.xml_sheet["A2"].value # 子订单号
        self.first_row = self.date_row(1) # 第一行数据列表
        self.max_row = self.xml_sheet.max_row # 最大行数
        self.max_column = self.xml_sheet.max_column # 最大列数


    def date_row(self, row):# 某一行的数据列表
        date_row_list = []
        for i in self.xml_sheet[str(row)]:
            date_row_list.append(i)
        return date_row_list


if __name__ == '__main__':
    a = ReadExcel("./圣萝莎erp输出输出参考.xlsx", "ERP表")
    # print(a.first_row())
    # for i in a.first_row():
    #     print(i.value)
    print(a.Eorder)
    print(a.first_row)
    print(a.max_row)
    print(a.max_column)